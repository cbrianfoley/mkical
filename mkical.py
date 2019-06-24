#!/usr/bin/python

#Reads a shared google spreadsheet, converts to ical file
import openpyxl
import io
import urllib2
import pytz
from icalendar import Calendar, Event
from datetime import datetime


gkey = '' #Insert GKEY here
sheetname = 'Schedule' #This is the name of the tab to parse
time = 3 #Default length of an event
tz = "America/Chicago" # Timezone


# Loads file
sheeturl = 'https://docs.google.com/spreadsheets/d/' + gkey + '/export?format=xlsx'
#print "Loading data from Google Spredsheet (%s)..." % (sheeturl)
wb = openpyxl.load_workbook(io.BytesIO(urllib2.urlopen(sheeturl).read()), data_only=True)

# cleanup_ws(ws) - cleans up the cell content in a worksheet so it is compatible with cue generation
# Strip leading/trailing whitespace on all string cells, and remove embedded newlines (thanks Brett)
# Remove instances of "
def cleanup_ws(ws):
   for row in ws.rows:
      for col,cell in enumerate(row):
         if isinstance(cell.value, basestring):
            cell.value = cell.value.strip() # get rid of leading/trailing whitespace
            cell.value = cell.value.replace('\n',' ') # get rid of embedded newlines
            cell.value = cell.value.replace('"', '\'') # get rid of double quotes (replace with single quotes)
            if cell.value == "": cell.value = None

# Verify that Schedule worksheet exists, and clean it
if sheetname in wb.get_sheet_names():
#   print "Found Schedule sheet in workbook!"
   sheet = wb[sheetname]
   cleanup_ws(sheet)
else: die("Sheet not found in workbook")

events = {}
data = []

#keys = sheet.iter_rows(min_row=1, max_row=1, max_col=3)
for row in sheet.iter_rows(min_row=3, max_col=2):
   for col,cell in enumerate(row):
      if col == 0:
         cell_data = str(row[col].value)
         cell_data = cell_data.replace(' ', '')
         cell_data = cell_data.replace('-','')
         cell_data = cell_data.replace(':','')
         events['Date']=cell_data
      else:
         cell_data = str(row[col].value)
         events['What']=cell_data
         data.append(events.copy()) 

cal = Calendar()

for row in data:
   event = Event()
   event.add('summary', row['What'])
   event.add('dtstart', datetime(int(row['Date'][:4]),int(row['Date'][4:6]),int(row['Date'][6:8]),int(row['Date'][8:10]),int(row['Date'][10:12]),0,tzinfo=pytz.timezone(tz)))
   event.add('dtend', datetime(int(row['Date'][:4]),int(row['Date'][4:6]),int(row['Date'][6:8]),int(row['Date'][8:10])+time,int(row['Date'][10:12]),0,tzinfo=pytz.timezone(tz)))
   event.add('description', row['What'])
   event.add('location', row['What'])
   cal.add_component(event)

cal = cal.to_ical()
f = open('.', 'wb')
f.write(cal)
f.close()
